"""
SACISC - Modulo 04 - Generador de Formato 7 oficial (Excel + PDF)
Se ejecuta dentro de un GitHub Action (repository_dispatch).
Toma un folio_id de pe_folios, jala el folio + sus 7 dias desde Supabase,
inyecta los valores en la plantilla oficial (SIN tocar formato/logo/imagenes),
genera el PDF desde ese mismo archivo, y sube ambos a Supabase Storage.
"""
import os
import sys
import json
import shutil
import subprocess
import requests
import openpyxl

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
FOLIO_ID = os.environ["FOLIO_ID"]

HEADERS = {
    "apikey": SUPABASE_SERVICE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    "Content-Type": "application/json",
}

TEMPLATE_PATH = "plantillas/FORMATO7_template.xlsx"
WORKDIR = "salida"
DIA_ROWS = {"L": 17, "M1": 18, "M2": 19, "J": 20, "V": 21, "S": 22, "D": 23}


def fetch_folio(folio_id):
    url = f"{SUPABASE_URL}/rest/v1/pe_folios"
    params = {"id": f"eq.{folio_id}", "select": "*,pe_folio_dias(*)"}
    resp = requests.get(url, headers=HEADERS, params=params)
    resp.raise_for_status()
    data = resp.json()
    if not data:
        raise ValueError(f"No se encontro el folio con id={folio_id}")
    return data[0]


def generar_excel(folio, out_path):
    shutil.copy(TEMPLATE_PATH, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["FORMATO 7"]

    # Encabezado
    ws["AD2"] = folio["folio"]
    ws["N6"] = folio["anio"]
    if folio.get("fecha_termino_semana"):
        y, m, d = folio["fecha_termino_semana"].split("-")
        ws["P6"] = int(m)
        ws["R6"] = int(d)
    ws["F8"] = "SERVICIOS CORPORATIVOS"
    ws["Q8"] = 75000
    ws["Y8"] = folio["subarea"]
    ws["F10"] = folio["nombre_trabajador"]
    ws["Y10"] = folio["ficha"]

    tot_dobletes, tot_comidas = 0, 0
    for dia in folio.get("pe_folio_dias", []):
        r = DIA_ROWS.get(dia["dia_semana"])
        if not r:
            continue
        ws[f"C{r}"] = dia.get("fecha")
        ws[f"D{r}"] = folio.get("salario")
        ws[f"E{r}"] = folio.get("jornada")
        ws[f"F{r}"] = folio.get("nivel")
        if folio["tipo_pago"] == "DOBLETE":
            if dia.get("horario") == "DESCANSO":
                ws[f"G{r}"] = "DESCANSO"
            elif dia.get("horario"):
                partes = dia["horario"].split("-")
                if len(partes) == 2:
                    ws[f"G{r}"] = partes[0]
                    ws[f"H{r}"] = partes[1]
            ws[f"I{r}"] = dia.get("dobletes") or None
        elif folio["tipo_pago"] == "TIEMPO_EXTRA":
            ws[f"K{r}"] = dia.get("horas") or None
            ws[f"L{r}"] = dia.get("minutos") or None
        elif folio["tipo_pago"] == "INSALUBRE":
            ws[f"Q{r}"] = dia.get("horas") or None
            ws[f"R{r}"] = dia.get("minutos") or None
        ws[f"U{r}"] = dia.get("comidas") or None
        ws[f"W{r}"] = folio.get("sociedad", "PMXC")
        ws[f"X{r}"] = folio.get("partida_presupuestal")
        ws[f"Y{r}"] = dia.get("labores_desarrolladas") or ""
        tot_dobletes += dia.get("dobletes") or 0
        tot_comidas += dia.get("comidas") or 0

    ws["I24"] = tot_dobletes or None
    ws["U24"] = tot_comidas or None

    # Pie: fecha y firmantes
    ws["B33"] = folio.get("elaboro_nombre")
    ws["M33"] = folio.get("vobo_nombre")
    ws["X31"] = folio.get("autoriza_nombre")

    wb.save(out_path)


def generar_pdf(xlsx_path, out_dir):
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "pdf", "--outdir", out_dir, xlsx_path],
        check=True,
    )


def subir_a_storage(local_path, bucket, dest_name):
    with open(local_path, "rb") as f:
        data = f.read()
    url = f"{SUPABASE_URL}/storage/v1/object/{bucket}/{dest_name}"
    headers = dict(HEADERS)
    headers["Content-Type"] = (
        "application/pdf" if dest_name.endswith(".pdf")
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    headers["x-upsert"] = "true"
    resp = requests.post(url, headers=headers, data=data)
    resp.raise_for_status()
    return f"{SUPABASE_URL}/storage/v1/object/public/{bucket}/{dest_name}"


def main():
    os.makedirs(WORKDIR, exist_ok=True)
    folio = fetch_folio(FOLIO_ID)

    base = f"FORMATO7_{folio['area']}_{folio['subarea']}_{folio['anio']}_{folio['folio']}".replace(" ", "_")
    xlsx_path = os.path.join(WORKDIR, f"{base}.xlsx")
    generar_excel(folio, xlsx_path)
    generar_pdf(xlsx_path, WORKDIR)
    pdf_path = os.path.join(WORKDIR, f"{base}.pdf")

    bucket = "formato7-generados"
    url_xlsx = subir_a_storage(xlsx_path, bucket, f"{base}.xlsx")
    url_pdf = subir_a_storage(pdf_path, bucket, f"{base}.pdf")

    print(json.dumps({"xlsx": url_xlsx, "pdf": url_pdf}))


if __name__ == "__main__":
    main()
